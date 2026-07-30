#!/usr/bin/env python
"""Construye las canastas (media y celiaca) FINALES y su Excel editable.

Decision de composicion (ver docs/CANASTAS.md): 15 items de alimentos y bebidas, elegidos por
consumo representativo (base CBA/INDEC), amplia cobertura regional (lat/lon) y estabilidad
temporal. 5 items tienen diferencial celiaco (harina, fideos, pan, galletitas dulces y saladas);
el resto es identico en ambas canastas.

Genera:
  config/canastas/canastas.xlsx  <- EDITABLE: modificar 'cantidad_mensual' a mano
  config/canastas/canastas.csv   <- export para git / lectura por el pipeline

Uso: python scripts/03_construir_canastas.py
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import duckdb
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
COV = ROOT / "data" / "processed" / "cobertura_productos_2026-06.parquet"

# grupo, cantidad_mensual (real, base CBA adulto-equivalente), unidad, tiene_gluten,
#        ean_media, ean_celiaco (== media si es naturalmente sin gluten)
CANASTA = [
    ("Harina",             1.35, "kg", 1, "7790070562258", "7792180142849"),
    ("Fideos",             1.00, "kg", 1, "7790070336118", "7790070335982"),
    ("Pan",                1.50, "kg", 1, "7793890258752", "7798221641616"),
    ("Galletitas dulces",  0.40, "kg", 1, "7622300742645", "7622202038099"),
    ("Galletitas saladas", 0.30, "kg", 1, "7795735000328", "7797330102377"),
    ("Arroz",              0.63, "kg", 0, "7790070433114", "7790070433114"),
    ("Aceite girasol",     1.20, "L",  0, "7790272001005", "7790272001005"),
    ("Azucar",             1.44, "kg", 0, "7792540250450", "7792540250450"),
    ("Yerba mate",         0.60, "kg", 0, "7793704000911", "7793704000911"),
    ("Cafe molido",        0.10, "kg", 0, "7790550000157", "7790550000157"),
    ("Leche entera",       7.35, "L",  0, "7790742363008", "7790742363008"),
    ("Atun",               0.30, "kg", 0, "7790580131364", "7790580131364"),
    ("Mermelada",          0.24, "kg", 0, "7790580132163", "7790580132163"),
    ("Lentejas",           0.24, "kg", 0, "7790580132453", "7790580132453"),
    ("Gaseosa cola",       4.00, "L",  0, "7790895000997", "7790895000997"),
]

_RE_SIZE = re.compile(r"(\d+[.,]?\d*)\s*(kg|kilo|grs|gr|g|ml|cc|lts|lt|l)\b")


def parse_envase(desc: str, unidad: str):
    """Extrae el tamano de envase de la descripcion, en la misma unidad (kg o L)."""
    matches = _RE_SIZE.findall((desc or "").lower())
    for val, u in reversed(matches):  # el ultimo suele ser el tamano del paquete
        val = float(val.replace(",", "."))
        if u in ("kg", "kilo"):        kg = val
        elif u in ("grs", "gr", "g"):  kg = val / 1000
        elif u in ("lts", "lt", "l"):  kg = val
        elif u in ("ml", "cc"):        kg = val / 1000
        else:                          continue
        return round(kg, 4)
    return None


def main() -> None:
    con = duckdb.connect()
    con.execute(f"CREATE VIEW cov AS SELECT * FROM '{COV.as_posix()}'")
    info = con.sql("SELECT id_producto, descripcion, marca, n_prov, n_suc, precio FROM cov").df()
    info = info.drop_duplicates("id_producto").set_index("id_producto")

    filas = []
    for grupo, cant, unidad, glut, ean_m, ean_c in CANASTA:
        m, c = info.loc[ean_m], info.loc[ean_c]
        env_m = parse_envase(m.descripcion, unidad)
        env_c = parse_envase(c.descripcion, unidad)
        filas.append(dict(
            grupo=grupo, cantidad_mensual=cant, unidad=unidad, tiene_gluten=glut,
            producto_media=m.descripcion, marca_media=m.marca, ean_media=ean_m,
            envase_media=env_m, precio_media=float(m.precio), cob_media=int(m.n_suc),
            producto_celiaco=c.descripcion, marca_celiaco=c.marca, ean_celiaco=ean_c,
            envase_celiaco=env_c, precio_celiaco=float(c.precio), cob_celiaco=int(c.n_suc),
        ))
    df = pd.DataFrame(filas)
    df["unidades_media"] = (df.cantidad_mensual / df.envase_media).round(2)
    df["unidades_celiaco"] = (df.cantidad_mensual / df.envase_celiaco).round(2)
    df["costo_media"] = (df.unidades_media * df.precio_media).round(0)
    df["costo_celiaco"] = (df.unidades_celiaco * df.precio_celiaco).round(0)

    cm, cc = df.costo_media.sum(), df.costo_celiaco.sum()
    print(df[["grupo", "cantidad_mensual", "unidad", "tiene_gluten",
              "costo_media", "costo_celiaco"]].to_string(index=False))
    print(f"\nCanasta MEDIA (2026-06):   ${cm:,.0f}")
    print(f"Canasta CELIACA (2026-06): ${cc:,.0f}")
    print(f"PRIMA CELIACA: {cc/cm-1:+.1%}")

    outdir = ROOT / "config" / "canastas"
    df.to_csv(outdir / "canastas.csv", index=False)
    escribir_excel(df, outdir / "canastas.xlsx", cm, cc)
    print(f"\nEscritos: {outdir/'canastas.csv'} y canastas.xlsx (editable)")


def escribir_excel(df: pd.DataFrame, path: Path, cm: float, cc: float) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Canasta"
    azul = PatternFill("solid", fgColor="0055A4"); amar = PatternFill("solid", fgColor="FFF2CC")
    blanco = Font(color="FFFFFF", bold=True)

    ws["A1"] = "Canasta de alimentos - editar solo la columna 'cantidad_mensual'"
    ws["A1"].font = Font(bold=True, size=12)
    cols = ["grupo", "cantidad_mensual", "unidad", "tiene_gluten",
            "producto_media", "marca_media", "ean_media", "envase_media",
            "producto_celiaco", "marca_celiaco", "ean_celiaco", "envase_celiaco"]
    ws.append([]); ws.append(cols)
    for j, _ in enumerate(cols, 1):
        c = ws.cell(row=3, column=j); c.fill = azul; c.font = blanco
    for _, r in df.iterrows():
        ws.append([r[c] for c in cols])
    # resaltar la columna editable (B)
    for i in range(4, 4 + len(df)):
        ws.cell(row=i, column=2).fill = amar
    for j in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(j)].width = max(
            12, min(46, df[cols[j-1]].astype(str).map(len).max() + 2 if cols[j-1] in df else 14))
    ws.freeze_panes = "A4"

    rs = wb.create_sheet("Resumen")
    rs["A1"] = "Resumen (precios 2026-06, se recalcula al correr el notebook)"
    rs["A1"].font = Font(bold=True)
    for i, (k, v) in enumerate([("Canasta media", f"${cm:,.0f}"),
                                ("Canasta celiaca", f"${cc:,.0f}"),
                                ("Prima celiaca", f"{cc/cm-1:+.1%}")], start=3):
        rs.cell(row=i, column=1, value=k); rs.cell(row=i, column=2, value=v)
    wb.save(path)


if __name__ == "__main__":
    main()
